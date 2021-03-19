import os
import pandas as pd
import pyautogui as pg
import time
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import subprocess


def process_exists(process_name):
    call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
    output = subprocess.check_output(call).decode()
    last_line = output.strip().split('\r\n')[-1]
    return last_line.lower().startswith(process_name.lower())


if not process_exists('Trebuchet.App.exe'):
    os.startfile('C:\Program Files (x86)\Cherwell Software 9.5.3\Cherwell Service Management 9.5.3\Trebuchet.App.exe')
    time.sleep(60)
else:
    pg.getWindowsWithTitle("Cherwell Service Management (licensed to Sutter Health)")[0].minimize()
    pg.getWindowsWithTitle("Cherwell Service Management (licensed to Sutter Health)")[0].maximize()
    time.sleep(1)


# Open the CAB Report
rzreportbutton = pg.locateOnScreen('screenshots\rzreport.png', confidence=0.8, region=(700, 70, 300, 50))
button_x, button_y = pg.center(rzreportbutton)
pg.click(button_x, button_y)
time.sleep(3)

# # Export the report as-is
print('Running Hotkeys')
time.sleep(5)
pg.hotkey('Alt', 'f')
pg.press(['Up', 'Up', 'Up', 'Up', 'Up', 'Up', 'Enter'])
pg.write('I:\CM\RZ.csv')
pg.press(['Tab', 'Tab', 'Tab', 'Tab', 'Tab', 'Tab', 'Space', 'Tab', 'Enter'])
pg.press('Enter')

# Pull CSV into DataFrame
print('Opening CSV')
time.sleep(5)
df = pd.read_csv('I:\CM\RZ.csv')

df = df.replace(r'\\r\\n', '\n', regex=True)

# Filter and Format DataFrame

thisReport = df[['CAB Review Notes', 'Change Coordinator Team', 'Status', 'Change Reason', 'Change ID', 'Proposed Start Date',
                 'Proposed End Date', 'Standard Template', 'Title', 'Owned By']]
thisReport.columns = ['Notes', 'Owned By Team', 'Disposition', 'Justification', 'Change #', 'Start', 'End', 'Standard Template',
                      'Title', 'Change Manager']

thisReport = thisReport.sort_values(by='Start')

# Create new XLSX file and write data frame to it
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(thisReport, index=False, header=True):
    ws.append(r)

# set the width of the columns
ws.column_dimensions['A'].width = 134/7
ws.column_dimensions['B'].width = 140/7
ws.column_dimensions['C'].width = 140/7
ws.column_dimensions['D'].width = 400/7
ws.column_dimensions['E'].width = 134/7
ws.column_dimensions['F'].width = 134/7
ws.column_dimensions['G'].width = 134/7
ws.column_dimensions['H'].width = 240/7
ws.column_dimensions['I'].width = 200/7
ws.column_dimensions['J'].width = 200/7
ws.row_dimensions[1].height = 30

# set wrap text to true and set alignment of the cells
z = y = 0
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        if z == 0:
            if y == 2:
                cell.fill = PatternFill(start_color='246194', end_color='246194', fill_type='solid')
            else:
                fillColor = '%02x%02x%02x' % (89, 89, 89)
                cell.fill = PatternFill(start_color=fillColor, end_color=fillColor, fill_type='solid')
            cell.font = Font(b=True, color="FFFFFF")
            y += 1
    z += 1

#Save WB and Open it
x = datetime.date.today()
print('Saving WB')
wb.save(f'I:\CM\RZ Reports\{x} RZ Review.xlsx')
print('Finished')
time.sleep(1)
subprocess.Popen([f'I:\\CM\\RZ Reports\\{x} RZ Review.xlsx'], shell=True)
