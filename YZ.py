import os
import pandas as pd
import pyautogui as pg
import time
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import subprocess


def process_exists(process_name):
    call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
    output = subprocess.check_output(call).decode()
    last_line = output.strip().split('\r\n')[-1]
    return last_line.lower().startswith(process_name.lower())


# Open Cherwell if it's not open, otherwise make it the focused screen
if not process_exists('Trebuchet.App.exe'):
    os.startfile('C:/Program Files (x86)/Cherwell Software 9.5.3/Cherwell Service Management 9.5.3/Trebuchet.App.exe')
    time.sleep(60)
else:
    pg.getWindowsWithTitle("Cherwell Service Management (licensed to Sutter Health)")[0].minimize()
    pg.getWindowsWithTitle("Cherwell Service Management (licensed to Sutter Health)")[0].maximize()
    time.sleep(1)

# # Open the YZ Report
yzreportbutton = pg.locateOnScreen('screenshots/yzreport.png', confidence=0.8, region=(700, 70, 300, 50))
button_x, button_y = pg.center(yzreportbutton)
pg.click(button_x, button_y)
time.sleep(3)

# Export the report as-is
print('Running Hotkeys')
time.sleep(5)
pg.hotkey('Alt', 'f')
pg.press(['Up', 'Up', 'Up', 'Up', 'Up', 'Up', 'Enter'])
pg.write('I:\\CM\\YZ.csv')
pg.press(['Tab', 'Tab', 'Tab', 'Tab', 'Tab', 'Tab', 'Space', 'Tab', 'Enter'])
pg.press('Enter')

# Pull CSV into DataFrame, filter and clean the data
print('Opening CSV')
time.sleep(5)
df = pd.read_csv('I:/CM/YZ.csv')
#df['Proposed Start Date'] = pd.to_datetime(df['Proposed Start Date'], format='%m/%d/%Y %I:%M %p').dt.strftime('%m/%d/%Y %H:%M')
#df['Proposed End Date'] = pd.to_datetime(df['Proposed End Date'], format='%m/%d/%Y %I:%M %p').dt.strftime('%m/%d/%Y %H:%M')
myfilter = df['CAB Review Notes'].str.contains('Need EC') | \
           df['CAB Review Notes'].str.contains('Needs EC') | \
           df['CAB Review Notes'].str.contains('NEED EC')
df = df.loc[myfilter]
statusCheck = df['Status'].str.contains('Implementation Approval') | \
              df['Status'].str.contains('Ready for Implementation')
df = df.loc[statusCheck]
df = df.replace(r'\\r\\n', '\n', regex=True)

# Format and Sort DataFrame
thisReport = df[['Status', 'Change Reason', 'Normal Change Level', 'Change ID', 'Proposed Start Date',
                 'Proposed End Date', 'Title', 'Change Coordinator Team', 'Owned By']]
thisReport.columns = ['Disposition', 'Justification', 'Level', 'Change #', 'Start', 'End', 'Title', 'Team',
                      'Change Manager']
thisReport = thisReport.sort_values(by='Start')

# Add Service column
blankTeams = [' ' for _ in range(len(thisReport.index))]
thisReport.insert(loc=2, column='Service', value=blankTeams)

# Initialize workbook and append data to it
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(thisReport, index=False, header=True):
    ws.append(r)

# set the width of the columns
ws.column_dimensions['A'].width = 134/7
ws.column_dimensions['B'].width = 360/7
ws.column_dimensions['C'].width = 64/7
ws.column_dimensions['D'].width = 104/7
ws.column_dimensions['E'].width = 104/7
ws.column_dimensions['F'].width = 104/7
ws.column_dimensions['G'].width = 104/7
ws.column_dimensions['H'].width = 185/7
ws.column_dimensions['I'].width = 155/7
ws.column_dimensions['J'].width = 155/7
ws.row_dimensions[1].height = 30

# set wrap text to true, set alignment of the cells, format first row as needed, and perform conditional formatting
z = y = 0
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        if z == 0:
            if y == 0:
                cell.fill = PatternFill(start_color='246194', end_color='246194', fill_type='solid')
                y += 1
            else:
                fillColor = '%02x%02x%02x' % (89, 89, 89)
                cell.fill = PatternFill(start_color=fillColor, end_color=fillColor, fill_type='solid')
            cell.font = Font(b=True, color="FFFFFF")
    z += 1

# Save WB and Open it
x = datetime.date.today()
wb.save(f'I:/CM/YZ Reports/{x} YZ Review.xlsx')
time.sleep(5)
subprocess.Popen([f'I:\\CM\\YZ Reports\\{x} YZ Review.xlsx'], shell=True)
