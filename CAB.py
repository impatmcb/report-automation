import os
import pandas as pd
import pyautogui as pg
import time
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import subprocess


def process_exists(process_name):
    call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
    output = subprocess.check_output(call).decode()
    last_line = output.strip().split('\r\n')[-1]
    return last_line.lower().startswith(process_name.lower())


if not process_exists('Trebuchet.App.exe'):
    os.startfile('C:\Program Files (x86)\Cherwell Software 9.5.3\Cherwell Service Management 9.5.3\Trebuchet.App.exe')
    time.sleep(60)
else:
    pg.getWindowsWithTitle("Cherwell Service Management (licensed to Sutter Health)")[0].minimize()
    pg.getWindowsWithTitle("Cherwell Service Management (licensed to Sutter Health)")[0].maximize()
    time.sleep(1)


# Open the CAB Report
cabreportbutton = pg.locateOnScreen('screenshots\cabreport.png', confidence=0.8, region=(700, 70, 300, 50))
button_x, button_y = pg.center(cabreportbutton)
pg.click(button_x, button_y)
time.sleep(3)

# Export the report as-is FINISHED
print('Running Hotkeys')
time.sleep(5)
pg.hotkey('Alt', 'f')
pg.press(['Up', 'Up', 'Up', 'Up', 'Up', 'Up', 'Enter'])
pg.write('I:\CM\CAB.csv')
pg.press(['Tab', 'Tab', 'Tab', 'Tab', 'Tab', 'Tab', 'Space', 'Tab', 'Enter'])
pg.press('Enter')

# Pull CSV into DataFrame
print('Opening CSV')
time.sleep(5)
df = pd.read_csv('I:\CM\CAB.csv')

# Find date for next CAB - NEED TO ADD HOLIDAY EXCLUSIONS
x = datetime.date.today()
if datetime.date.today().weekday() < 2:
    while True:
        x += datetime.timedelta(days=1)
        if x.weekday() == 2:
            nextCab = x.strftime('%#m/%d/%Y')
            break
else:
    while True:
        x += datetime.timedelta(days=1)
        if x.weekday() == 0:
            nextCab = x.strftime('%#m/%d/%Y')
            break

# Filter and Format DataFrame
targetDate = df['Taget Date CAB'] == nextCab
thisReport = df[targetDate]
thisReport = thisReport[['Change Coordinator', 'Owned By', 'Change ID', 'Proposed Start Date', 'Proposed End Date',
                         'Title', 'Change Coordinator Team']]
thisReport.columns = ['Coordinator', 'Manager', 'Change', 'Start', 'End', 'Title', 'Team']
thisReport['Start'] = pd.to_datetime(thisReport['Start']).dt.strftime('%m/%d/%Y %H:%M')
thisReport['End'] = pd.to_datetime(thisReport['End']).dt.strftime('%m/%d/%Y %H:%M')
thisReport = thisReport.sort_values(by='Start')

# Create new XLSX file and write data frame to it
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(thisReport, index=False, header=True):
    ws.append(r)

# set the width of the columns
ws.column_dimensions['A'].width = 150/7
ws.column_dimensions['B'].width = 150/7
ws.column_dimensions['C'].width = 60/7
ws.column_dimensions['D'].width = 130/7
ws.column_dimensions['E'].width = 130/7
ws.column_dimensions['F'].width = 240/7
ws.column_dimensions['G'].width = 330/7
ws.row_dimensions[1].height = 30

# set header style formatting, set wrap text to true, and set alignment of the cells
z = 0
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        if z == 0:
            cell.fill = PatternFill(start_color='246194', end_color='246194', fill_type='solid')
            cell.font = Font(b=True, color="FFFFFF")
    z += 1


#Save WB and Open it
print('Saving WB')
try:
    wb.save(f'I:\CM\CAB Reports\{x} CAB Review.xlsx')
    print('Finished')
    time.sleep(1)
    subprocess.Popen([f'I:\\CM\\CAB Reports\\{x} Cab Review.xlsx'], shell=True)
except PermissionError:
    print("Close the report and run the script again...")
